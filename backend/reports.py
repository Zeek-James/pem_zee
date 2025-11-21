"""
Report generation module for Excel and PDF exports
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import config


class ReportGenerator:
    """Generate reports in Excel and PDF formats"""

    def __init__(self):
        # Ensure reports directory exists
        os.makedirs(config.REPORTS_DIR, exist_ok=True)

    def generate_excel_report(self, harvests, milling_records, storage_records, sales, report_type='summary'):
        """Generate comprehensive Excel report"""
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets
        if report_type == 'summary' or report_type == 'all':
            self._create_harvest_sheet(wb, harvests)
            self._create_milling_sheet(wb, milling_records)
            self._create_storage_sheet(wb, storage_records)
            self._create_sales_sheet(wb, sales)
            self._create_summary_sheet(wb, harvests, milling_records, storage_records, sales)
        elif report_type == 'harvest':
            self._create_harvest_sheet(wb, harvests)
        elif report_type == 'milling':
            self._create_milling_sheet(wb, milling_records)
        elif report_type == 'storage':
            self._create_storage_sheet(wb, storage_records)
        elif report_type == 'sales':
            self._create_sales_sheet(wb, sales)

        # Save file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'palm_oil_report_{report_type}_{timestamp}.xlsx'
        filepath = os.path.join(config.REPORTS_DIR, filename)
        wb.save(filepath)

        return filepath

    def _create_harvest_sheet(self, wb, harvests):
        """Create harvest data sheet"""
        ws = wb.create_sheet('Harvest Records')

        # Headers
        headers = ['ID', 'Date', 'Plantation', 'Bunches', 'Weight/Bunch (kg)',
                   'Total Weight (kg)', 'Ripeness', 'Expected Yield (kg)']
        ws.append(headers)

        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for h in harvests:
            ws.append([
                h.id,
                h.harvest_date.strftime('%Y-%m-%d'),
                h.plantation,
                h.num_bunches,
                h.weight_per_bunch,
                h.total_weight,
                h.ripeness,
                h.expected_oil_yield
            ])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_milling_sheet(self, wb, milling_records):
        """Create milling data sheet"""
        ws = wb.create_sheet('Milling Records')

        headers = ['ID', 'Date', 'Mill Location', 'Harvest ID', 'Milling Cost (₦)',
                   'Transport Cost (₦)', 'Oil Yield (kg)', 'Cost/kg (₦)', 'Total Cost (₦)']
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        for m in milling_records:
            ws.append([
                m.id,
                m.milling_date.strftime('%Y-%m-%d'),
                m.mill_location,
                m.harvest_id,
                m.milling_cost,
                m.transport_cost,
                m.oil_yield,
                round(m.cost_per_kg, 2),
                m.total_cost
            ])

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_storage_sheet(self, wb, storage_records):
        """Create storage data sheet"""
        ws = wb.create_sheet('Storage Inventory')

        headers = ['Container ID', 'Quantity (kg)', 'Storage Date', 'Expiry Date',
                   'Days Until Expiry', 'Plantation Source', 'Status']
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        for s in storage_records:
            status = 'Sold' if s.is_sold else ('Expired' if s.is_expired else 'Available')
            ws.append([
                s.container_id,
                s.quantity,
                s.storage_date.strftime('%Y-%m-%d'),
                s.expiry_date.strftime('%Y-%m-%d'),
                s.days_until_expiry,
                s.plantation_source,
                status
            ])

            # Highlight expired items
            if s.is_expired and not s.is_sold:
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    cell.font = Font(color='FFFFFF')

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_sales_sheet(self, wb, sales):
        """Create sales data sheet"""
        ws = wb.create_sheet('Sales Records')

        headers = ['ID', 'Date', 'Buyer', 'Container ID', 'Quantity (kg)',
                   'Price/kg (₦)', 'Total Revenue (₦)', 'Payment Status', 'Payment Date']
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        for s in sales:
            ws.append([
                s.id,
                s.sale_date.strftime('%Y-%m-%d'),
                s.buyer_name,
                s.storage.container_id if s.storage else 'N/A',
                s.quantity_sold,
                s.price_per_kg,
                s.total_revenue,
                s.payment_status,
                s.payment_date.strftime('%Y-%m-%d') if s.payment_date else 'N/A'
            ])

            # Highlight pending payments
            if s.is_payment_pending:
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_summary_sheet(self, wb, harvests, milling_records, storage_records, sales):
        """Create summary sheet with KPIs"""
        ws = wb.create_sheet('Summary', 0)

        # Title
        ws['A1'] = 'PALM OIL BUSINESS SUMMARY'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')

        # Date
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:D2')

        # KPIs
        row = 4
        ws[f'A{row}'] = 'KEY PERFORMANCE INDICATORS'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 2

        # Calculate KPIs
        total_ffb = sum(h.total_weight for h in harvests)
        total_oil = sum(m.oil_yield for m in milling_records)
        total_cost = sum(m.total_cost for m in milling_records)
        total_revenue = sum(s.total_revenue for s in sales)
        total_profit = total_revenue - total_cost
        available_stock = sum(s.quantity for s in storage_records if not s.is_sold)

        kpis = [
            ['Total FFB Harvested', f'{total_ffb:.2f} kg'],
            ['Total CPO Produced', f'{total_oil:.2f} kg'],
            ['Total Production Cost', f'₦{total_cost:,.2f}'],
            ['Total Revenue', f'₦{total_revenue:,.2f}'],
            ['Total Profit', f'₦{total_profit:,.2f}'],
            ['Available Stock', f'{available_stock:.2f} kg'],
            ['Number of Harvests', len(harvests)],
            ['Number of Milling Operations', len(milling_records)],
            ['Number of Sales', len(sales)]
        ]

        for kpi, value in kpis:
            ws[f'A{row}'] = kpi
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def generate_pdf_report(self, harvests, milling_records, storage_records, sales, report_type='summary'):
        """Generate PDF report"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'palm_oil_report_{report_type}_{timestamp}.pdf'
        filepath = os.path.join(config.REPORTS_DIR, filename)

        doc = SimpleDocTemplate(filepath, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=30,
            alignment=1  # Center
        )

        elements.append(Paragraph('PALM OIL BUSINESS REPORT', title_style))
        elements.append(Paragraph(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', styles['Normal']))
        elements.append(Spacer(1, 20))

        # Summary
        total_ffb = sum(h.total_weight for h in harvests)
        total_oil = sum(m.oil_yield for m in milling_records)
        total_cost = sum(m.total_cost for m in milling_records)
        total_revenue = sum(s.total_revenue for s in sales)
        total_profit = total_revenue - total_cost

        summary_data = [
            ['Metric', 'Value'],
            ['Total FFB Harvested', f'{total_ffb:.2f} kg'],
            ['Total CPO Produced', f'{total_oil:.2f} kg'],
            ['Total Production Cost', f'₦{total_cost:,.2f}'],
            ['Total Revenue', f'₦{total_revenue:,.2f}'],
            ['Total Profit', f'₦{total_profit:,.2f}']
        ]

        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(Paragraph('SUMMARY', styles['Heading2']))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Add detailed sections based on report type
        if report_type in ['summary', 'all', 'harvest']:
            elements.append(PageBreak())
            elements.append(Paragraph('HARVEST RECORDS', styles['Heading2']))
            elements.append(Spacer(1, 10))

            harvest_data = [['ID', 'Date', 'Plantation', 'Bunches', 'Total (kg)']]
            for h in harvests[:20]:  # Limit to 20 records for PDF
                harvest_data.append([
                    str(h.id),
                    h.harvest_date.strftime('%Y-%m-%d'),
                    h.plantation,
                    str(h.num_bunches),
                    f'{h.total_weight:.2f}'
                ])

            harvest_table = Table(harvest_data, colWidths=[0.5*inch, 1*inch, 1*inch, 0.8*inch, 1*inch])
            harvest_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#70AD47')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            elements.append(harvest_table)

        doc.build(elements)
        return filepath
